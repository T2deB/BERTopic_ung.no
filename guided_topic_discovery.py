#!/usr/bin/env python3
"""
guided_topic_discovery.py — Iterative residual c-TF-IDF per category.

For each category in the data, runs sequential passes of c-TF-IDF on bigrams.
Each pass removes questions containing the dominant bigrams from the previous
pass, revealing the next most coherent theme underneath.

Output format per category:
    Pass 1: top bigrams (+ % questions containing each)
    Pass 2: top bigrams after removing pass-1 questions
    Pass 3: top bigrams after removing pass-1 and pass-2 questions
    ...

Also writes an Excel file with one sheet per category showing the
columnar format: [Pass 1 bigrams | Pass 2 bigrams | Pass 3 bigrams ...]

Usage:
    python guided_topic_discovery.py --segment girls_13_15
    python guided_topic_discovery.py --segment girls_13_15 --category kropp
    python guided_topic_discovery.py --segment all
"""

import argparse
import json
import re
import warnings
from collections import Counter
from pathlib import Path
from typing import List, Tuple, Dict

import numpy as np
import pandas as pd

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("WARNING: openpyxl not installed — Excel output will be skipped. pip install openpyxl")

try:
    from nltk.corpus import stopwords as _nltk_stop
    NLTK_STOP = set(_nltk_stop.words("norwegian"))
except Exception:
    NLTK_STOP = set()

SCRIPT_DIR     = Path(__file__).parent
DATA_DIR       = r"C:\Users\TorsteinDeBesche\NIFU\21571 Folkehelse og livsmestring - General\Ap2c Informasjonskanal for unge\Positron_project\data"
OUTPUT_DIR     = r"C:\Users\TorsteinDeBesche\NIFU\21571 Folkehelse og livsmestring - General\Ap2c Informasjonskanal for unge\Positron_project\outputs"
SEGMENT_SUBDIR = "segments"
SEGMENTS       = ["boys_13_15", "boys_16_20", "girls_13_15", "girls_16_20"]

# Discovery settings
N_PASSES           = 5     # number of iterative passes per category
TOP_BIGRAMS_SHOWN  = 10    # bigrams to display per pass
TOP_BIGRAMS_REMOVE = 3     # top TF-IDF bigrams shown as pass label
TOP_FREQ_REMOVE    = 5     # top FREQUENCY bigrams used for actual removal
MIN_BIGRAM_FREQ    = 0.005  # minimum fraction of questions (0.5%) — closer to R's n>=5
MIN_BIGRAM_ABS     = 5      # absolute minimum for both display and removal
# Note: no relative threshold for removal — absolute count scales across
# categories of all sizes. A 5-question floor works for 500 or 50,000 questions.
MIN_QUESTIONS      = 50    # minimum questions remaining to continue passes

# Matches R script's custom_stop
CUSTOM_STOP = {
    "hei", "hallo", "takk", "hilsen",
    "jeg", "du", "vi", "dere", "han", "hun", "de", "den", "det", "dette",
    "ikke", "bare", "litt", "veldig", "også", "så", "å",
    "må", "kan", "vil", "skal", "burde",
    "er", "var", "blir", "ble", "har", "hadde",
    "som", "til", "for", "på", "av", "i", "om", "fra", "med", "uten",
    "når", "hva", "hvordan", "hvor", "hvorfor",
}

# Extra stopwords for bigrams only — matches R script's glue_words
# Prevents meaningless bigrams like "følelser for", "snakke om"
GLUE_WORDS = {
    "for", "om", "til", "at", "som", "å", "med", "uten", "hos", "fra",
    "på", "av", "i", "den", "det", "de", "dette", "da", "nå", "så",
    "jeg", "du", "vi", "han", "hun",
    "er", "var", "blir", "ble", "har", "hadde",
    "ikke", "bare", "veldig", "litt",
}

# Bigrams to always exclude — matches R script's ban_bigrams
BAN_BIGRAMS = {
    "hele tiden", "hele tatt", "flere ganger", "lang tid", "svar fort", "gammel jente", "gammel gutt",
}

STOP_WORDS = sorted(NLTK_STOP | CUSTOM_STOP)


# ── Topic column join ────────────────────────────────────────────────────────

def _norm_body(series: pd.Series) -> pd.Series:
    """Normalise body text for joining — must match bottom_up_clustering.py."""
    s = (
        series.astype(str)
        .str.lower()
        .str.strip()
        .str.replace(r"\s+", " ", regex=True)
    )
    s = s.str.replace(
        r"^(hei|heisann|hallo|halloi)[\s,!:.\-]+", "", regex=True
    ).str.strip()
    return s


def build_topic_lookup(data_dir: Path) -> pd.DataFrame:
    """
    Scan DATA_DIR for *.csv files and build a (body_key, topic) lookup.
    Normalises body text the same way as the age join so keys match.
    Returns DataFrame with columns [body_key, topic].
    """
    csv_files = sorted(
        p for p in data_dir.glob("*.csv")
        if p.suffix.lower() == ".csv" and p.stat().st_size > 0
    )
    if not csv_files:
        warnings.warn(f"[topic] No CSV files found in {data_dir}")
        return pd.DataFrame(columns=["body_key", "topic"])

    # Detect text column from first file
    text_col = None
    try:
        header = pd.read_csv(csv_files[0], nrows=0, encoding="utf-8-sig",
                             sep=None, engine="python")
        for candidate in ["body", "question", "text", "body_norm", "sporsmal", "tekst"]:
            if candidate in header.columns:
                text_col = candidate
                break
    except Exception as exc:
        warnings.warn(f"[topic] Could not read header: {exc}")
        return pd.DataFrame(columns=["body_key", "topic"])

    if text_col is None:
        warnings.warn("[topic] Could not identify text column")
        return pd.DataFrame(columns=["body_key", "topic"])

    try:
        header2 = pd.read_csv(csv_files[0], nrows=0, encoding="utf-8-sig",
                              sep=None, engine="python")
        if "topic" not in header2.columns:
            warnings.warn(f"[topic] No 'topic' column found. Columns: {header2.columns.tolist()}")
            return pd.DataFrame(columns=["body_key", "topic"])
    except Exception:
        pass

    parts = []
    for p in csv_files:
        try:
            chunk = pd.read_csv(
                p, usecols=[text_col, "topic"],
                dtype={text_col: str, "topic": str},
                encoding="utf-8-sig",
                low_memory=True,
            )
            parts.append(chunk)
        except Exception as exc:
            warnings.warn(f"[topic] Could not read {p.name}: {exc}")

    if not parts:
        return pd.DataFrame(columns=["body_key", "topic"])

    raw = pd.concat(parts, ignore_index=True)
    raw["body_key"] = _norm_body(raw[text_col])

    # Split comma-separated topics into separate rows (matches R separate_rows)
    # e.g. "Sex, Kropp" becomes two rows: one for "Sex", one for "Kropp"
    raw["topic"] = raw["topic"].astype(str).str.split(",")
    raw = raw.explode("topic")
    raw["topic"] = raw["topic"].str.strip()
    raw = raw[raw["topic"].str.len() > 0]
    raw = raw[raw["topic"] != "-"]
    raw = raw[raw["topic"] != "nan"]

    # Keep all (body_key, topic) pairs — one question can appear in multiple categories
    lookup = (
        raw[["body_key", "topic"]]
        .dropna(subset=["body_key", "topic"])
        .drop_duplicates(subset=["body_key", "topic"])
    )
    print(f"[topic] lookup built: {len(lookup):,} rows from {len(parts)} files")
    print(f"[topic] categories found: {sorted(lookup['topic'].unique().tolist())}")
    return lookup


def join_topic_to_assignments(assign: pd.DataFrame, data_dir: Path) -> pd.DataFrame:
    """Join topic column onto assignments dataframe via normalised body text."""
    lookup = build_topic_lookup(data_dir)
    if lookup.empty:
        print("[topic] WARNING: empty lookup — topic column not joined")
        assign = assign.copy()
        assign["topic"] = "unknown"
        return assign

    body_keys = _norm_body(assign["body_norm"])
    assign = assign.copy()
    assign["body_key"] = body_keys.values

    # Merge — a question with multiple topics produces multiple rows
    assign_expanded = assign.merge(lookup, on="body_key", how="left")
    assign_expanded["topic"] = assign_expanded["topic"].fillna("unknown")

    coverage = (assign_expanded["topic"] != "unknown").mean()
    print(f"[topic] join coverage: {coverage:.1%}")
    print(f"[topic] questions after expansion: {len(assign_expanded):,} "
          f"(was {len(assign):,} — {len(assign_expanded)-len(assign):,} extra rows from multi-category questions)")
    print(f"[topic] category distribution:")
    for cat, cnt in assign_expanded["topic"].value_counts().items():
        print(f"         {cat:<35} {cnt:>6,}")

    assign_expanded = assign_expanded.drop(columns=["body_key"])
    return assign_expanded


# ── c-TF-IDF ────────────────────────────────────────────────────────────────

def _clean_text(text: str) -> str:
    """Match R script clean_text: lowercase, strip URLs, numbers, punctuation."""
    text = text.lower()
    text = re.sub(r'http\S+|www\S+', ' ', text)
    text = re.sub(r'[0-9]+', ' ', text)
    text = re.sub(r'[^\w\sæøå]', ' ', text)   # keep letters incl Norwegian
    text = re.sub(r'\s+', ' ', text).strip()
    return text


def _tokenize_bigrams(text: str) -> List[str]:
    """
    Tokenize into bigrams matching R script logic:
    - both words length >= 3
    - neither word in STOP_WORDS or GLUE_WORDS
    - result bigram not in BAN_BIGRAMS
    """
    words = text.split()
    words = [w for w in words if len(w) >= 3]
    bigrams = []
    for i in range(len(words) - 1):
        w1, w2 = words[i], words[i + 1]
        if (w1 not in STOP_WORDS and w2 not in STOP_WORDS
                and w1 not in GLUE_WORDS and w2 not in GLUE_WORDS):
            bg = f"{w1} {w2}"
            if bg not in BAN_BIGRAMS:
                bigrams.append(bg)
    return bigrams


def compute_ctfidf_bigrams(
    texts: List[str],
    top_k: int = TOP_BIGRAMS_SHOWN,
    min_freq: float = MIN_BIGRAM_FREQ,
) -> List[Tuple[str, float, float]]:
    """
    Compute c-TF-IDF for bigrams.
    Text cleaning matches R script: lowercase, strip URLs/numbers/punctuation.
    Bigram filtering matches R script: glue_words removed, ban_bigrams excluded.

    Returns list of (bigram, ctfidf_score, freq_pct).
    Only returns bigrams appearing in >= min_freq of questions.
    """
    n = len(texts)
    if n < 10:
        return []

    # Texts are expected to already be cleaned by the caller (discover_category).
    # If called directly with raw texts, clean here as fallback.
    doc_bigrams = [_tokenize_bigrams(t) for t in texts]

    min_df = max(2, int(n * min_freq * 0.5))
    doc_freq: Counter = Counter()
    for bg_list in doc_bigrams:
        for bg in set(bg_list):
            doc_freq[bg] += 1

    vocab = [bg for bg, df in doc_freq.items() if df >= min_df]
    if not vocab:
        return []
    vocab_idx = {bg: i for i, bg in enumerate(vocab)}
    V = len(vocab)

    tf_agg   = np.zeros(V, dtype=np.float64)
    doc_flag = np.zeros(V, dtype=np.float64)
    for bg_list in doc_bigrams:
        seen: set = set()
        for bg in bg_list:
            if bg in vocab_idx:
                tf_agg[vocab_idx[bg]] += 1
                if bg not in seen:
                    doc_flag[vocab_idx[bg]] += 1
                    seen.add(bg)

    total = tf_agg.sum()
    if total == 0:
        return []
    tf_norm = tf_agg / total

    idf = np.log((n + 1) / (1 + doc_flag)) + 1.0
    scores = tf_norm * idf
    freq   = doc_flag / n

    abs_counts = doc_flag  # already computed above
    mask = (freq >= min_freq) & (abs_counts >= MIN_BIGRAM_ABS)
    if not mask.any():
        return []

    filtered_scores = scores * mask
    top_idx = np.argsort(-filtered_scores)[:top_k * 2]
    top_idx = [i for i in top_idx if mask[i]][:top_k]

    return [(vocab[i], float(scores[i]), float(freq[i])) for i in top_idx]


def questions_containing_bigrams(texts: List[str], bigrams: List[str]) -> np.ndarray:
    """
    Returns boolean array: True for each question whose token bigram list
    contains at least one of the given bigrams.

    Uses _tokenize_bigrams (same as scoring) so that short connector words
    like 'og', 'i' are filtered before adjacency is checked.
    This means 'mamma og pappa' correctly matches the bigram 'mamma pappa'
    because tokenization skips 'og' (< 3 chars) making 'mamma' and 'pappa'
    adjacent in the token list.
    """
    bigram_set = set(bigrams)
    mask = np.zeros(len(texts), dtype=bool)
    for i, text in enumerate(texts):
        doc_bigrams = _tokenize_bigrams(text)
        if any(bg in bigram_set for bg in doc_bigrams):
            mask[i] = True
    return mask


# ── Main discovery loop ──────────────────────────────────────────────────────

def compute_ctfidf_bigrams_cross_category(
    category_texts: Dict[str, List[str]],
    top_k: int = TOP_BIGRAMS_SHOWN,
    min_freq: float = MIN_BIGRAM_FREQ,
) -> Dict[str, List[Tuple[str, float, float]]]:
    """
    Cross-category c-TF-IDF for bigrams.
    Treats each category as a document (matches R's bind_tf_idf approach).
    IDF = log((n_cats + 1) / (n_cats_containing_bigram + 1)) + 1 —
    bigrams appearing in many categories score lower, highlighting
    what is distinctive to each category.

    Args:
        category_texts: dict mapping category name -> list of pre-cleaned texts
        top_k:          top bigrams to return per category
        min_freq:       minimum fraction of a category's questions that must
                        contain the bigram (applied after IDF weighting)

    Returns:
        dict mapping category name -> list of (bigram, ctfidf_score, freq_pct)
    """
    cats = list(category_texts.keys())
    n_cats = len(cats)
    if n_cats == 0:
        return {}

    # Per-category bigram counts
    cat_tf: Dict[str, Counter] = {}        # bigram -> total occurrences in category
    cat_doc_flag: Dict[str, Counter] = {}  # bigram -> n questions containing bigram
    cat_n: Dict[str, int] = {}

    for cat, texts in category_texts.items():
        cat_n[cat] = len(texts)
        tf_c: Counter = Counter()
        flag_c: Counter = Counter()
        for text in texts:
            bgs = _tokenize_bigrams(text)
            for bg in bgs:
                tf_c[bg] += 1
            for bg in set(bgs):
                flag_c[bg] += 1
        cat_tf[cat] = tf_c
        cat_doc_flag[cat] = flag_c

    # Cross-category document frequency: how many categories contain each bigram
    cross_doc_freq: Counter = Counter()
    for cat in cats:
        for bg in cat_doc_flag[cat]:
            cross_doc_freq[bg] += 1

    results: Dict[str, List[Tuple[str, float, float]]] = {}

    for cat in cats:
        n = cat_n[cat]
        if n < 10:
            results[cat] = []
            continue

        tf_counter   = cat_tf[cat]
        flag_counter = cat_doc_flag[cat]
        min_df = max(2, int(n * min_freq * 0.5))
        vocab  = [bg for bg, df in flag_counter.items() if df >= min_df]
        if not vocab:
            results[cat] = []
            continue

        total_tf = sum(tf_counter[bg] for bg in vocab)
        if total_tf == 0:
            results[cat] = []
            continue

        scored = []
        for bg in vocab:
            tf   = tf_counter[bg] / total_tf
            # Match R's bind_tf_idf: log(n_docs / df)
            # Bigrams appearing in ALL categories get idf=0 → score=0
            # This prevents generic cross-category bigrams like "mamma pappa"
            # from appearing in category-specific results
            df = cross_doc_freq[bg]
            idf = np.log(n_cats / df) if df < n_cats else 0.0
            freq  = flag_counter[bg] / n
            n_q   = flag_counter[bg]  # raw question count
            score = tf * idf
            n_q = flag_counter[bg]
            if freq >= min_freq and n_q >= MIN_BIGRAM_ABS and score > 0:
                scored.append((bg, float(score), float(freq), int(n_q)))

        scored.sort(key=lambda x: -x[1])
        results[cat] = scored[:top_k]

    return results


def get_top_freq_bigrams(
    texts: List[str],
    top_k: int = TOP_FREQ_REMOVE,
) -> List[Tuple[str, float, int]]:
    """
    Get the most FREQUENT bigrams in texts regardless of TF-IDF score.
    Used for removal so each pass covers maximum questions.
    Returns list of (bigram, freq_pct, n_questions) sorted by frequency desc.
    """
    n = len(texts)
    if n < 10:
        return []

    doc_freq: Counter = Counter()
    for text in texts:
        bgs = _tokenize_bigrams(text)
        for bg in set(bgs):
            doc_freq[bg] += 1

    results = []
    for bg, count in doc_freq.items():
        freq = count / n
        if count >= MIN_BIGRAM_ABS:  # absolute threshold only — scales across corpus sizes
            results.append((bg, float(freq), int(count)))

    results.sort(key=lambda x: -x[1])
    return results[:top_k]


def compute_cooccurrence(
    texts: List[str],
    tfidf_bigrams: List[tuple],
    freq_bigrams: List[Tuple[str, float, int]],
) -> Dict[str, Dict[str, float]]:
    """
    For each frequency bigram (used for removal), compute what fraction
    of the questions containing it also contain each TF-IDF bigram.

    Returns dict: freq_bigram -> {tfidf_bigram -> overlap_pct}
    Only reports overlaps >= 10%.

    Interpretation: if removing "mamma pappa" takes 40% of "tik tok" questions,
    those themes are entangled. If overlap is 5%, removal is clean.
    """
    if not tfidf_bigrams or not freq_bigrams:
        return {}

    tfidf_terms = [entry[0] for entry in tfidf_bigrams]
    freq_terms  = [bg for bg, _, _ in freq_bigrams]

    # Pre-compute bigram sets per document
    doc_bigram_sets = [set(_tokenize_bigrams(t)) for t in texts]

    result: Dict[str, Dict[str, float]] = {}

    for fb in freq_terms:
        fb_docs = [i for i, bgs in enumerate(doc_bigram_sets) if fb in bgs]
        if not fb_docs:
            continue

        overlaps: Dict[str, float] = {}
        for tb in tfidf_terms:
            if tb == fb:
                continue
            both = sum(1 for i in fb_docs if tb in doc_bigram_sets[i])
            pct  = both / len(fb_docs)
            if pct >= 0.10:
                overlaps[tb] = round(pct, 3)

        if overlaps:
            result[fb] = dict(sorted(overlaps.items(), key=lambda x: -x[1]))

    return result


def discover_category_cross(
    texts: List[str],
    category: str,
    all_category_texts: Dict[str, List[str]],
    n_passes: int = N_PASSES,
) -> List[Dict]:
    """
    Iterative residual c-TF-IDF using cross-category IDF for display,
    frequency-based removal for coverage.

    Each pass:
    1. Computes cross-category TF-IDF → shows distinctive bigrams (interpretable)
    2. Computes within-category frequency → removes most common bigrams (coverage)

    This separates the two goals:
    - What is this topic about?  → answered by TF-IDF bigrams displayed
    - Which questions to remove? → answered by frequency bigrams removed
    """
    cleaned_texts = [_clean_text(t) for t in texts]
    remaining = list(range(len(cleaned_texts)))
    results = []

    for pass_num in range(1, n_passes + 1):
        n_remaining   = len(remaining)
        pct_remaining = n_remaining / len(cleaned_texts)

        if n_remaining < MIN_QUESTIONS:
            print(f"    Pass {pass_num}: stopping — only {n_remaining} questions left")
            break

        pass_texts = [cleaned_texts[i] for i in remaining]

        # --- DISPLAY: cross-category TF-IDF bigrams (distinctive vocabulary) ---
        cross_ctx           = dict(all_category_texts)
        cross_ctx[category] = pass_texts
        cross_results       = compute_ctfidf_bigrams_cross_category(
            cross_ctx, top_k=TOP_BIGRAMS_SHOWN, min_freq=MIN_BIGRAM_FREQ
        )
        tfidf_bigrams = cross_results.get(category, [])

        # --- REMOVAL: top frequency bigrams (maximum coverage) ---
        freq_bigrams = get_top_freq_bigrams(pass_texts, top_k=TOP_FREQ_REMOVE)
        remove_bigrams = [b for b, _, _ in freq_bigrams]

        # Fallback: if no bigrams meet frequency threshold, use TF-IDF top bigrams
        if not remove_bigrams and tfidf_bigrams:
            remove_bigrams = [b for b, _, _, *_ in tfidf_bigrams[:TOP_BIGRAMS_REMOVE]]

        if not tfidf_bigrams and not remove_bigrams:
            print(f"    Pass {pass_num}: stopping — no bigrams above threshold")
            break

        # Compute co-occurrence between TF-IDF and frequency bigrams
        cooc = compute_cooccurrence(pass_texts, tfidf_bigrams, freq_bigrams)

        results.append({
            "pass_num":        pass_num,
            "n_questions":     n_remaining,
            "pct_remaining":   round(pct_remaining, 3),
            "bigrams":         tfidf_bigrams,
            "removed_bigrams": remove_bigrams,
            "freq_bigrams":    [(b, f, n) for b, f, n in freq_bigrams],
            "cooccurrence":    cooc,
        })

        # Print TF-IDF bigrams (what this topic is about)
        if tfidf_bigrams:
            print(f"    Pass {pass_num} ({n_remaining:,} q, {pct_remaining:.0%} remaining):")
            for entry in tfidf_bigrams[:TOP_BIGRAMS_SHOWN]:
                n_q = entry[3] if len(entry) > 3 else int(entry[2] * n_remaining)
                print(f"      [tfidf] {entry[0]:<32} {entry[2]*100:>5.1f}%  ({n_q:,} q)")
        else:
            print(f"    Pass {pass_num} ({n_remaining:,} q): no distinctive TF-IDF bigrams")

        # Print frequency bigrams (what gets removed)
        if freq_bigrams:
            print(f"      Top frequency bigrams (used for removal):")
            for bg, freq, n_q in freq_bigrams[:TOP_FREQ_REMOVE]:
                print(f"        [freq] {bg:<32} {freq*100:>5.1f}%  ({n_q:,} q)")

        # Print co-occurrence where meaningful
        if cooc:
            print(f"      Co-occurrence (freq bigram → overlapping TF-IDF bigrams):")
            for fb, overlaps in cooc.items():
                overlap_str = ", ".join(
                    f"{tb} ({v*100:.0f}%)" for tb, v in list(overlaps.items())[:3]
                )
                print(f"        [{fb}] also contains: {overlap_str}")

        # Remove questions containing the frequency bigrams
        remove_mask = questions_containing_bigrams(pass_texts, remove_bigrams)
        n_removed   = int(remove_mask.sum())
        keep_local  = np.where(~remove_mask)[0]
        remaining   = [remaining[i] for i in keep_local]

        print(f"      → removed {n_removed:,} questions "
              f"({n_removed/n_remaining*100:.1f}% of remaining) "
              f"containing: {', '.join(remove_bigrams[:3])}{'...' if len(remove_bigrams) > 3 else ''}")
        print()

    return results


def discover_category(
    texts: List[str],
    category: str,
    n_passes: int = N_PASSES,
) -> List[Dict]:
    """
    Run n_passes of iterative residual c-TF-IDF on texts.
    Texts are cleaned once upfront so bigram detection and removal
    operate on identical text — fixing the mismatch where bigrams
    were computed on cleaned text but removal searched raw text.
    """
    # Clean once upfront — both scoring and removal use the same cleaned texts
    cleaned_texts = [_clean_text(t) for t in texts]

    remaining = list(range(len(cleaned_texts)))
    results = []

    for pass_num in range(1, n_passes + 1):
        n_remaining = len(remaining)
        pct_remaining = n_remaining / len(cleaned_texts)

        if n_remaining < MIN_QUESTIONS:
            print(f"    Pass {pass_num}: stopping — only {n_remaining} questions left")
            break

        pass_texts = [cleaned_texts[i] for i in remaining]
        bigrams = compute_ctfidf_bigrams(pass_texts, top_k=TOP_BIGRAMS_SHOWN)

        if not bigrams:
            print(f"    Pass {pass_num}: stopping — no bigrams above threshold")
            break

        remove_bigrams = [entry[0] for entry in bigrams[:TOP_BIGRAMS_REMOVE]]

        results.append({
            "pass_num": pass_num,
            "n_questions": n_remaining,
            "pct_remaining": round(pct_remaining, 3),
            "bigrams": bigrams,
            "removed_bigrams": remove_bigrams,
        })

        print(f"    Pass {pass_num} ({n_remaining:,} q, {pct_remaining:.0%} remaining):")
        for bigram, score, freq, n_q in bigrams[:TOP_BIGRAMS_SHOWN]:
            print(f"      {bigram:<35} {freq*100:>5.1f}%  ({n_q:,} q)")

        # Remove questions containing the top bigrams — using cleaned texts
        remove_mask = questions_containing_bigrams(pass_texts, remove_bigrams)
        n_removed = int(remove_mask.sum())
        keep_local = np.where(~remove_mask)[0]
        remaining = [remaining[i] for i in keep_local]

        print(f"      → removed {n_removed:,} questions containing: "
              f"{', '.join(remove_bigrams)}")
        print()

    return results


def run_discovery(seg_name: str, target_category: str = None) -> None:
    seg_dir = Path(OUTPUT_DIR) / SEGMENT_SUBDIR / seg_name
    parquet  = seg_dir / "A_cluster_assignments.parquet"

    if not parquet.exists():
        print(f"ERROR: no assignments parquet at {parquet}")
        return

    print(f"\n{'='*60}")
    print(f"Segment: {seg_name}")
    print(f"{'='*60}")

    assign = pd.read_parquet(parquet)
    print(f"Loaded {len(assign):,} questions")

    assign = join_topic_to_assignments(assign, Path(DATA_DIR))

    categories = (
        [target_category] if target_category
        else sorted(assign["topic"].dropna().unique().tolist())
    )
    categories = [c for c in categories if c and c != "unknown"]

    # Pre-clean all category texts once for cross-category IDF
    print(f"\nPre-cleaning texts for {len(categories)} categories...")
    all_category_texts: Dict[str, List[str]] = {}
    for cat in categories:
        cat_df = assign[assign["topic"] == cat]
        texts  = cat_df["body_norm"].astype(str).tolist()
        if len(texts) >= MIN_QUESTIONS:
            all_category_texts[cat] = [_clean_text(t) for t in texts]

    all_results: Dict[str, List[Dict]] = {}

    for cat in categories:
        cat_df = assign[assign["topic"] == cat]
        texts  = cat_df["body_norm"].astype(str).tolist()
        print(f"\n{'─'*50}")
        print(f"Category: {cat}  ({len(texts):,} questions)")
        print(f"{'─'*50}")

        if len(texts) < MIN_QUESTIONS:
            print(f"  Too few questions, skipping.")
            continue

        results = discover_category_cross(texts, cat, all_category_texts)
        all_results[cat] = results

    _write_text_report(all_results, seg_name, seg_dir)
    if HAS_OPENPYXL:
        _write_excel_report(all_results, seg_name, seg_dir)
    _write_json(all_results, seg_name, seg_dir)

    print(f"\n[done] Discovery complete for {seg_name}")


# ── Output writers ───────────────────────────────────────────────────────────

def _write_text_report(
    all_results: Dict[str, List[Dict]],
    seg_name: str,
    seg_dir: Path,
) -> None:
    lines = [f"Guided topic discovery — {seg_name}", "=" * 80, ""]

    for cat, passes in all_results.items():
        if not passes:
            continue
        lines.append(f"{cat.upper()}")
        lines.append("─" * 60)

        header_parts = []
        for p in passes:
            label = f"Pass {p['pass_num']} ({p['n_questions']:,} q)"
            if p['pass_num'] > 1:
                prev = passes[p['pass_num'] - 2]
                label += f" [removed: {', '.join(prev['removed_bigrams'])}]"
            header_parts.append(f"{label:<40}")
        lines.append("  ".join(header_parts))
        lines.append("")

        max_rows = max(len(p["bigrams"]) for p in passes)
        for row_i in range(max_rows):
            row_parts = []
            for p in passes:
                if row_i < len(p["bigrams"]):
                    entry  = p["bigrams"][row_i]
                    bigram, _, freq = entry[0], entry[1], entry[2]
                    n_q = entry[3] if len(entry) > 3 else ""
                    cell = f"{bigram}  ({freq*100:.1f}%, {n_q:,} q)" if n_q else f"{bigram}  ({freq*100:.1f}%)"
                else:
                    cell = ""
                row_parts.append(f"{cell:<45}")
            lines.append("  ".join(row_parts))

        lines.append("")
        lines.append("")

    out_path = seg_dir / f"topic_discovery_{seg_name}.txt"
    out_path.write_text("\n".join(lines), encoding="utf-8")
    print(f"\nText report written: {out_path.name}")


def _write_excel_report(
    all_results: Dict[str, List[Dict]],
    seg_name: str,
    seg_dir: Path,
) -> None:
    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    HDR_FILL   = PatternFill("solid", fgColor="2C3E50")
    HDR_FONT   = Font(color="FFFFFF", bold=True, size=11)
    CAT_FILL   = PatternFill("solid", fgColor="ECF0F1")
    CAT_FONT   = Font(bold=True, size=12)
    PASS_FILLS = [
        PatternFill("solid", fgColor="D6EAF8"),
        PatternFill("solid", fgColor="D5F5E3"),
        PatternFill("solid", fgColor="FDEBD0"),
        PatternFill("solid", fgColor="F9EBEA"),
        PatternFill("solid", fgColor="EAF2FF"),
    ]
    thin   = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cat, passes in all_results.items():
        if not passes:
            continue

        # Excel sheet names cannot contain: / \ ? * [ ] :
        _INVALID = r'[/\\?*\[\]:]'
        sheet_name = re.sub(_INVALID, '-', cat).strip()[:31]
        ws = wb.create_sheet(title=sheet_name)

        # Category title spanning all columns
        ws.merge_cells(start_row=1, start_column=1,
                       end_row=1, end_column=len(passes) * 2)
        cell = ws.cell(row=1, column=1, value=cat.upper())
        cell.font = CAT_FONT
        cell.fill = CAT_FILL
        cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[1].height = 22

        # Pass headers
        for p_idx, p in enumerate(passes):
            col = p_idx * 2 + 1
            pass_label = f"Pass {p['pass_num']}"
            sub_label  = f"{p['n_questions']:,} q  ({p['pct_remaining']*100:.0f}% remaining)"
            if p['pass_num'] > 1:
                prev_removed = passes[p['pass_num'] - 2]["removed_bigrams"]
                sub_label += f"\nRemoved: {', '.join(prev_removed)}"

            ws.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col + 1)
            c = ws.cell(row=2, column=col, value=pass_label)
            c.font = HDR_FONT
            c.fill = HDR_FILL
            c.alignment = Alignment(horizontal="center")

            removed_label = f"Removed: {', '.join(passes[p_idx-1]['removed_bigrams'][:2])}..." if p_idx > 0 and passes[p_idx-1].get('removed_bigrams') else "Pass 1 (full corpus)"
            for cc, label in [(col, f"TF-IDF bigrams\n({removed_label})"), (col + 1, "% (n questions)")]:
                cell = ws.cell(row=3, column=cc, value=label)
                cell.font = Font(bold=True, size=10)
                cell.fill = PASS_FILLS[p_idx % len(PASS_FILLS)]
                cell.alignment = Alignment(horizontal="center")

        # Data rows
        max_rows = max(len(p["bigrams"]) for p in passes)
        for row_i in range(max_rows):
            excel_row = row_i + 4
            for p_idx, p in enumerate(passes):
                col  = p_idx * 2 + 1
                fill = PASS_FILLS[p_idx % len(PASS_FILLS)]
                if row_i < len(p["bigrams"]):
                    entry  = p["bigrams"][row_i]
                    bigram = entry[0]
                    freq   = entry[2]
                    n_q    = entry[3] if len(entry) > 3 else None
                    c1 = ws.cell(row=excel_row, column=col,     value=bigram)
                    pct_label = f"{round(freq*100, 1)}% ({n_q:,})" if n_q else f"{round(freq*100, 1)}%"
                    c2 = ws.cell(row=excel_row, column=col + 1, value=pct_label)
                    for c in [c1, c2]:
                        c.fill = fill
                        c.border = border
                        c.alignment = Alignment(vertical="center")
                else:
                    for cc in [col, col + 1]:
                        ws.cell(row=excel_row, column=cc).fill = fill

        # Co-occurrence sections (appended below data rows per pass)
        base_cooc_row = max_rows + 4 + 1  # one row below the last data row
        for p_idx, p in enumerate(passes):
            p_cooc = p.get("cooccurrence", {})
            if not p_cooc:
                continue
            col  = p_idx * 2 + 1
            fill = PASS_FILLS[p_idx % len(PASS_FILLS)]
            next_row = base_cooc_row
            cooc_label = ws.cell(
                row=next_row, column=col,
                value="Co-occurrence (freq → TF-IDF)"
            )
            cooc_label.font = Font(italic=True, size=10, color="888780")
            cooc_label.fill = fill
            next_row += 1
            for fb, overlaps in list(p_cooc.items())[:3]:
                overlap_str = ", ".join(
                    f"{tb} {v*100:.0f}%"
                    for tb, v in list(overlaps.items())[:2]
                )
                cooc_cell = ws.cell(
                    row=next_row, column=col,
                    value=f"{fb} → {overlap_str}"
                )
                cooc_cell.font = Font(italic=True, size=10)
                cooc_cell.fill = fill
                ws.merge_cells(
                    start_row=next_row, start_column=col,
                    end_row=next_row, end_column=col + 1
                )
                next_row += 1

        # Column widths
        for p_idx in range(len(passes)):
            col = p_idx * 2 + 1
            ws.column_dimensions[get_column_letter(col)].width     = 32
            ws.column_dimensions[get_column_letter(col + 1)].width = 14

        ws.freeze_panes = ws.cell(row=4, column=1)

    out_path = seg_dir / f"topic_discovery_{seg_name}.xlsx"
    wb.save(str(out_path))
    print(f"Excel report written: {out_path.name}")


def _write_json(
    all_results: Dict[str, List[Dict]],
    seg_name: str,
    seg_dir: Path,
) -> None:
    def _clean(obj):
        if isinstance(obj, np.integer):  return int(obj)
        if isinstance(obj, np.floating): return float(obj)
        if isinstance(obj, dict):        return {k: _clean(v) for k, v in obj.items()}
        if isinstance(obj, (list, tuple)): return [_clean(v) for v in obj]
        return obj

    out_path = seg_dir / f"topic_discovery_{seg_name}.json"
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(_clean(all_results), f, ensure_ascii=False, indent=2)
    print(f"JSON written: {out_path.name}")


# ── Entry point ──────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Iterative residual c-TF-IDF per category.")
    parser.add_argument("--segment",  default="girls_13_15",
                        help="Segment name or 'all'")
    parser.add_argument("--category", default=None,
                        help="Single category slug, e.g. 'kropp'. Default: all.")
    args = parser.parse_args()

    segs = SEGMENTS if args.segment == "all" else [args.segment]
    for seg in segs:
        try:
            run_discovery(seg, args.category)
        except Exception as e:
            print(f"ERROR in {seg}: {e}")
            import traceback
            traceback.print_exc()


if __name__ == "__main__":
    main()
